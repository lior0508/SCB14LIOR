'''
接口自动化步骤：
1、excel 测试用例准备ok，代码自动读取测试数据
2、发送接口请求，得到相应信息   ---api1()
3、断言：实际结果 --- 预期结果  通过/不通过
4、写入通过、不通过  ----excel---write_result

第三方库：操作excel表格----openpyxl库：实现读取excel测试数据，并写入
1.安装pip install openpyxl    2. 导入库

注意：方便读取数据，excel表格放到跟python文件同一级
excel中的三大对象：1.工作簿workbook 2.sheet  3.cell

'''
#定义读取函数
import openpyxl
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook('test_case_api.xlsx')#加载工作簿--文档名字
    sheet = wb['register']# 获取表单
    max_row = sheet.max_row #----获取最大行数---15行
    #max_column = sheet.max_column          获取最大列数
    #print(max_row,max_column)
    case_list = [] #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(       #把信息转换成字典形式
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, #2-1通过表单获取行号，列号获取单元格的数据，2-5 获取url
        data = sheet.cell(row=i,column=6).value,   # 获取data
        expect = sheet.cell(row=i,column=7).value   #获取expect
        )
        case_list.append(dict1)  #每循环一次，就把读取的字典数据存放在这个list上面
    #print(case_list)
    #print(url,data,expect)
    return (case_list)  #返回测试用例列表
# cases = read_data('test_case_api.xlsx','register')      #调用函数传参，返回测试用例结果赋值给一个定义case变量
# print(cases)

#写入结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result #写入结果
    wb.save(filename)   #保存结果，，记得先关闭文档
# write_result('test_case_api.xlsx','login',3,8,'Failed')

import requests
# 接口函数
def api1(url,data):
    url_login = 'http://120.78.128.25:8766/futureloan/member/login'   #接口地址
    headers_login = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  #请求头{} 里面输入具体的请求头内容
    data_login = {"mobile_phone":15989548730,"pwd":"lemon666"}   #请求正文
    #request.post(url=url_login,json=data_login,headers=headers_login)
    result = requests.post(url=url,json=data,headers=headers_login) #接受post方法的结果
    # print(result.text)   #res只是简单的反馈，具体需要再加约束来获取
    # print(result.json())
    response =result.json()
    return response

#执行测试用例并会写实际结果
def execute_fun(filename,sheetname,):
    cases = read_data('test_case_api.xlsx','register')  #调用读取测试用例，获取所有的测试用例在，并定义一个变量接收
    # 这里获取的是
    #print(cases)
    for case in cases:
        case_id = case['case_id']   #case.get('case_id')另一种方式
        url = case.get('url')
        data = eval(case.get('data'))  #   print(type(data))为字符串，需要转换成字典---eval()可以直接把括号里面的内容去掉引号再运行
        expect =eval(case.get('expect'))   #获取预期结果,和data一样，从表格里获取过来都带有引号--字符串的格式
        expect_msg = expect.get('msg')   # [1]也可以
        real_result = api1(url=url,data=data) #调用发送接口请求函数，返回结果，用变量来接收
        real_msg = real_result.get('msg')
        print('预期结果中的msg:{}'.format(expect_msg))
        print('预期结果中的msg:{}'.format(real_msg))
        if real_msg == expect_msg:
            print("第{}条测试用例执行通过".format(case_id))
            final_result = "passed"
        else:
            print('第{}条测试用例执行不通过'.format(case_id))
            final_result = 'failed'
        write_result('test_case_api.xlsx','register',case_id+1,8, final_result)
        print('*'*20)