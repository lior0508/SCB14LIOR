# list0 = [3,3.14,'lior',[1,3,5,7]]  #变量取值尽量不要用到已有的函数名
# print(list0)
# print(list0[2])
# print(list0[0:3])
# print(list0[:len(list0)])
# print(list0[3][2]) #嵌套取值
# # #
# # # #增
# # # list0.append('贺雄')
# # # print(list0) #----[3, 3.14, 'lior', [1, 3, 5, 7], '贺雄']-----追加在元素列表末尾
# # # list0.insert(1,"希望")
# # # print(list0)  # [3, '希望', 3.14, 'lior', [1, 3, 5, 7], '贺雄']  ,指定元素位置插入
# # # list0.append('贺雄')
# # # print(list0) #[3, '希望', 3.14, 'lior', [1, 3, 5, 7], '贺雄', '贺雄']  :元素可以重复存在
# # #
# # # #删
# # # #list0.pop()  #默认删除最后一个元素,()不需要输入
# # # list0.pop(1)
# # # print(list0)  # [3, 3.14, 'lior', [1, 3, 5, 7], '贺雄', '贺雄']   指定位置删除
# # #
# # # list0.remove("贺雄")
# # # print(list0)    #[3, 3.14, 'lior', [1, 3, 5, 7], '贺雄']  按元素来删除一个
# # #
# # # # list0.clear()  #清空列表元素
# # # # print(list0)
# # #
# # # #改
# # # list0[1]= 6.28  #找到元素，直接赋值
# # # print(list0)  # [3, 6.28, 'lior', [1, 3, 5, 7], '贺雄']
# # #
# # # # 元组：tuple()  了解一下
# # # # 1、元素组成：任意数据类型
# # # # 2、元组取值：索引0开始，参考字符片取值
# # # # 3、tuple元素是不可变的，
# # # tuple1 = (3,3.14,'lior',[1,3,5,7])
# # # print(tuple1)
# # # print(tuple1[2])
# # #
# # # # 字典：dict{} 键值对--key：value--（元素）
# # # # 1、使用场景：描述一个对象的基本信息：属性一值
# # # #2、key：除了字典和列表以外都可取（不可变得数据类型）--一般用来定义后续元素的意思；value：元素--可以任意数据类型
# # # # 3、字面没有顺序--没有索引，，取值---取value值--变量名，
# dict1 = {'name':'lior','age':20,'hobby':'study','sex':'男'}
# print(dict1) #  {'name': 'lior', 'age': 20, 'hobby': 'study', 'sex': '男'}
# print(dict1['age'])   # 20   通过key取值
# print(dict1.values()) # dict_values(['lior', 20, 'study', '男'])
# print(dict1.get("age"))   #study
# print(dict1.items())   # dict_items([('name', 'lior'), ('age', 20), ('hobby', 'study'), ('sex', '男')])
# print(dict1.values())
# # #
# # # #增
# # # dict1['city'] = '深圳'   #key 不存在，新增键值对
# # # print(dict1)  # {'name': 'lior', 'age': 20, 'hobby': 'study', 'sex': '男', 'city': '深圳'}
# # # dict1.update({'hight':175,'weight':100})   #可以插入一个或者多个值
# # # print(dict1)  #{'name': 'lior', 'age': 20, 'hobby': 'study', 'sex': '男', 'city': '深圳', 'hight': 175, 'weight': 100}
# # #
# # # #改
# # # dict1['weight'] = 130     #key 存在，更新value
# # # print(dict1)   #{'name': 'lior', 'age': 20, 'hobby': 'study', 'sex': '男', 'city': '深圳', 'hight': 175, 'weight': 130}
# # #
# # # #删
# # # dict1.pop('age')
# # # print(dict1)   #{'name': 'lior', 'hobby': 'study', 'sex': '男', 'city': '深圳', 'hight': 175, 'weight': 130}
# # # #字典没有顺序。指定key进行删除
# # #
# # # # python控制流：控制代码的走向，常用的2中：判断、循环
# # # #if 判断：语法
# # # '''
# # # if条件a：  ---条件为真
# # #   满足条件a，执行语句
# # # elif 条件b：         -----可以是多个条件，也可没有
# # # ...
# # # else:
# # #   不满足条件，执行语句
# # # '''
# # # money = 1000
# # # if money >=1000000:
# # #     print('国外游')
# # # elif money >= 100000:
# # #     print('国内游')
# # # elif money >=10000:
# # #     print('市内游')
# # # else:
# # #     print('好好搬砖吧')
# # #
# # # money = int(input('请输入您目前的金额:'))  # 强制转换成int数据类型的，之前是字符串类型
# # # if money >=1000000:
# # #     print('国外游')
# # # elif money >= 100000:
# # #     print('国内游')
# # # elif money >=10000:
# # #     print('市内游')
# # # else:
# # #     print('好好搬砖吧')
# # # # 请输入您目前的金额:100 按enter出现下列
# # # # 好好搬砖吧
# # #
# # # #for 循环：遍历这个数据对象(i字符串，列表，元组，字典)的每一个元素从头到尾访问一下
# # # #for  in   ： 固定用法
# # # #循环次数？  ----元素个数
# # # #循环结束?-----遍历结束。循环结束
# # # #中断循环？-----break 跳转结束整个循环，continue  跳转结束当前循环，继续下一个循环
# # # count = 0
# # # str1='我希望拿到一份高薪的工作'
# # # for i in str1:
# # #     #if i== '拿':
# # #      #break
# # #      #continue
# # #     print(i)
# # #     count += 1
# # #     print(count)
# # #     print('*'*15)
# # # print(count)   #循环总次数
# # # print(len(str1)) #元素个数
# # #
# # # #range (起始值，终点值，步长)：生成有一个整数序列，起始默认值为0，步长默认1
# # # #range(5) -----0、1、2、3、4
# # # for i in range(5):
# # #     print(i)
# # #     print(('*')*5)
# # #
# # # a = [1,2,"6",'summer']
# # # print('i' not in a)
# #
# # # dict_1={'class_id':45,'num':20}
# # # num =20
# # # if num >= 5:
# # #     print(dict_1['num'])
# # # list3['lior','小米椒','lucia','k','建文']
# # # dict_inf={'name':'lior'}
# # # dict_inf.update({'sex':'m','age':22,'city':'SZ'})
# # # print(dict_inf)
# # # info = {'name': 'lior', 'sex': 'm', 'age': 22, 'city': 'SZ'}
# # # info1=[]
# # # info1.append(info)
# # # print(info1)
# # # key = list(dict_inf.keys())
# # # print(key)
# # # key1=['name', 'sex', 'age', 'city']
# # # value = list(dict_inf.values())
# # # print(value)
# # # value1= ['lior', 'm', 22, 'SZ']
# # # d=[]
# # # for i in range(len(key1)):
# # #     d[key1[i]]=value1[i]
# # #     print(d)
# #
# def good_job(salary,bonus,subsidy,*args,**kwargs):
#     # salary = 8000
#     # bonus = 1000
#     # subsidy = 500
#     print('salary:{}'.format(salary))
#     print('bonus:{}'.format(bonus))
#     print('subsidy:{}'.format(subsidy))
#     print('args:{}'.format(args))
#     print('kwargs:{}'.format(kwargs))
#     suml = salary + bonus + subsidy
#     for i in args:
#         suml += i
#     for j in kwargs:
#         suml += kwargs.get(j)#kwargs[j]这种方法也可以
#     print('工资总和:{}'.format(suml))
#     return suml
# result=good_job(800,1000,500,100,200,a=150,b=150)
# print(result)
# if result >= 10000:
#     print('好工作')
# else:
#     print('就这工作')
# a = (1,3,5,7,9)
# b = list(a)
# print(b)
#
# A= (1,3,5,7,9)
# count = 0
# for i in A:
#     count += i
# print(count)
#
# a = {2,5,3,1,9,8}
# b = len(a)
# if b >= 5:
#     print("true")
# else:
#     print('false')
# print(b)
# dict1 = {'name':'lior','age':20}
# dict2 = dict(name='柯基',age=22)
# print(dict1)
# print(dict2)
# #二种方式
#
# '''

# list1=[1,2,3]
# list1.append(4)
# list2=list1
# print(list1)
# print(list2)
import openpyxl
def inf(filename,sheetname,row,column,name_re,age_re,sex_re):
#     name = input("请输入姓名:")
#     age = input("请输入年龄:")
#     sex = input("请输入性别:")
#     print('''
#     姓名:{}
#     性别:{}
#     年龄:{}
#     '''.format(name,sex,age))
#     print('我的名字{},今年{}岁，性别{}，喜欢敲代码'.format(name,age,sex))
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = name_re
    sheet.cell(row=row, column=column+1).value = age_re
    sheet.cell(row=row, column=column+2).value = sex_re
read_result = inf('inf.xlsx','sheet1',2,1,"lior",18,'m')

# str1 = 'python hello aaa 123123aabb'
# count = 0
# for i in str1:
#     if i == 'a':
#       count +=1
# print(count)
#
# print(str1.find('123'))
#
# info = ('name=name,age=age,sex=sex').value
# print(info)